"""
Integration Test

데이터 변환 파이프라인 통합 테스트
"""

import pytest
from datetime import time

from app.api.routes import convert_to_itinerary_days, parse_time, parse_region
from app.services.excel import ItineraryDay, ItineraryItem


class TestDataConversion:
    """데이터 변환 통합 테스트"""

    def test_parse_time_valid(self):
        """시간 파싱 테스트"""
        assert parse_time("09:00") == time(9, 0, 0)
        assert parse_time("14:30") == time(14, 30, 0)
        assert parse_time("23:59") == time(23, 59, 0)

    def test_parse_time_invalid(self):
        """잘못된 시간 파싱"""
        assert parse_time("") is None
        assert parse_time("invalid") is None
        assert parse_time("25:00") is None

    def test_parse_region_with_arrow(self):
        """화살표가 있는 지역 파싱"""
        from_region, to_region = parse_region("인천 → 오사카")
        assert from_region == "인천"
        assert to_region == "오사카"

        from_region, to_region = parse_region("서울->부산")
        assert from_region == "서울"
        assert to_region == "부산"

    def test_parse_region_without_arrow(self):
        """화살표가 없는 지역 파싱"""
        from_region, to_region = parse_region("교토")
        assert from_region == "교토"
        assert to_region == ""

    def test_convert_to_itinerary_days_single_day(self):
        """단일 일차 변환 테스트"""
        schedule_data = [
            {
                "day": "1일차",
                "date": "03/15(일)",
                "region": "인천 → 오사카",
                "transport": "항공",
                "time": "09:00",
                "schedule": "인천공항 출발",
                "description": "",
                "image_url": "",
                "meals": "조:기내식"
            }
        ]

        result = convert_to_itinerary_days(schedule_data)

        assert len(result) == 1
        assert result[0].day_number == 1
        assert len(result[0].items) == 1

        item = result[0].items[0]
        assert item.region_from == "인천"
        assert item.region_to == "오사카"
        assert item.transport == "항공"
        assert item.time == time(9, 0, 0)
        assert item.schedule_title == "인천공항 출발"
        assert item.description == ""
        assert item.meal == "조:기내식"

    def test_convert_to_itinerary_days_multiple_days(self):
        """다일 일차 변환 테스트"""
        schedule_data = [
            {
                "day": "1일차",
                "date": "03/15(일)",
                "region": "인천",
                "transport": "항공",
                "time": "09:00",
                "schedule": "출발",
                "description": "",
                "image_url": "",
                "meals": "조:기내식"
            },
            {
                "day": "1일차",
                "date": "03/15(일)",
                "region": "오사카",
                "transport": "전용차량",
                "time": "14:00",
                "schedule": "오사카성 관광",
                "description": "",
                "image_url": "",
                "meals": ""
            },
            {
                "day": "2일차",
                "date": "03/16(월)",
                "region": "교토",
                "transport": "전용차량",
                "time": "10:00",
                "schedule": "금각사 관광",
                "description": "",
                "image_url": "",
                "meals": "조:호텔식"
            }
        ]

        result = convert_to_itinerary_days(schedule_data)

        assert len(result) == 2

        # 1일차 검증
        assert result[0].day_number == 1
        assert len(result[0].items) == 2

        # 2일차 검증
        assert result[1].day_number == 2
        assert len(result[1].items) == 1

    def test_convert_to_itinerary_days_missing_fields(self):
        """필드 누락 시 기본값 처리"""
        schedule_data = [
            {
                "day": "1일차",
                "date": "",
                "region": "",
                "transport": "",
                "time": "",
                "schedule": "일정 상세",
                "description": "",
                "image_url": "",
                "meals": ""
            }
        ]

        result = convert_to_itinerary_days(schedule_data)

        assert len(result) == 1
        item = result[0].items[0]
        assert item.region_from == ""
        assert item.region_to == ""
        assert item.transport == ""
        assert item.time is None
        assert item.schedule_title == "일정 상세"
        assert item.description == ""
        assert item.meal == ""

    def test_convert_to_itinerary_days_day_grouping(self):
        """같은 일차 그룹화 테스트"""
        schedule_data = [
            {"day": "1일차", "date": "", "region": "A", "transport": "", "time": "", "schedule": "일정1", "description": "", "image_url": "", "meals": ""},
            {"day": "1일차", "date": "", "region": "B", "transport": "", "time": "", "schedule": "일정2", "description": "", "image_url": "", "meals": ""},
            {"day": "1일차", "date": "", "region": "C", "transport": "", "time": "", "schedule": "일정3", "description": "", "image_url": "", "meals": ""},
        ]

        result = convert_to_itinerary_days(schedule_data)

        assert len(result) == 1
        assert result[0].day_number == 1
        assert len(result[0].items) == 3
        assert result[0].items[0].schedule_title == "일정1"
        assert result[0].items[1].schedule_title == "일정2"
        assert result[0].items[2].schedule_title == "일정3"

    def test_convert_to_itinerary_days_day_order(self):
        """일차 순서 정렬 테스트"""
        schedule_data = [
            {"day": "3일차", "date": "", "region": "C", "transport": "", "time": "", "schedule": "3일차", "description": "", "image_url": "", "meals": ""},
            {"day": "1일차", "date": "", "region": "A", "transport": "", "time": "", "schedule": "1일차", "description": "", "image_url": "", "meals": ""},
            {"day": "2일차", "date": "", "region": "B", "transport": "", "time": "", "schedule": "2일차", "description": "", "image_url": "", "meals": ""},
        ]

        result = convert_to_itinerary_days(schedule_data)

        assert len(result) == 3
        assert result[0].day_number == 1
        assert result[1].day_number == 2
        assert result[2].day_number == 3

    def test_empty_schedule_data(self):
        """빈 스케줄 데이터"""
        result = convert_to_itinerary_days([])
        assert result == []


class TestEndToEndFlow:
    """End-to-End 데이터 흐름 테스트"""

    def test_scraper_to_excel_pipeline(self):
        """스크래퍼 -> 변환 -> 엑셀 파이프라인"""
        from app.services.excel import ExcelGenerator

        # 1. Mock 스크래퍼 데이터
        scraper_output = [
            {
                "day": "1일차",
                "date": "03/15(일)",
                "region": "인천 → 오사카",
                "transport": "KE123",
                "time": "09:00",
                "schedule": "인천공항 출발 → 오사카 간사이공항 도착",
                "description": "",
                "image_url": "",
                "meals": "조:기내식, 석:호텔식"
            },
            {
                "day": "2일차",
                "date": "03/16(월)",
                "region": "교토",
                "transport": "전용차량",
                "time": "10:00",
                "schedule": "금각사 관광",
                "description": "교토의 대표적인 사찰",
                "image_url": "https://image.hanatour.com/usr/cms/resize/800_0/photo/kinkakuji.jpg",
                "meals": "조:호텔식, 중:현지식"
            }
        ]

        # 2. 데이터 변환
        itinerary_days = convert_to_itinerary_days(scraper_output)

        # 3. 엑셀 생성
        generator = ExcelGenerator()
        excel_bytes = generator.generate_excel_bytes(itinerary_days)

        # 4. 검증
        assert excel_bytes is not None
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # 엑셀 파일 검증
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # 헤더 확인
        assert ws.cell(1, 1).value == "일자"
        assert ws.cell(1, 2).value == "지역"

        # 데이터 확인 (컬럼: 일자/지역/관광지/이미지/설명)
        assert ws.cell(2, 1).value == "1일차"
        assert ws.cell(2, 2).value == "인천"
        assert ws.cell(2, 3).value == "인천공항 출발 → 오사카 간사이공항 도착"
