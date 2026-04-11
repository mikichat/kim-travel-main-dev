"""
Excel Generator Service

하나투어 일정 데이터를 엑셀 파일로 변환하는 서비스
"""

import io
import os
import re
import urllib.request
from dataclasses import dataclass, field
from datetime import time
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class ItineraryItem:
    """일정 항목 데이터 모델"""

    region_from: str = ""  # 출발 지역
    region_to: str = ""  # 도착 지역
    transport: str = ""  # 교통편 (항공편명, 전용차량 등)
    time: Optional[time] = None  # 출발/도착 시간
    schedule_title: str = ""  # 관광지 이름
    description: str = ""  # 관광지 상세 설명
    image_url: str = ""  # 이미지 URL
    meal: str = ""  # 식사 정보 (조:중:석)


@dataclass
class ItineraryDay:
    """일자별 일정 데이터 모델"""

    day_number: int  # 일자 (1, 2, 3, ...)
    items: List[ItineraryItem]  # 일정 항목 리스트


class ExcelGenerator:
    """엑셀 일정표 생성기"""

    def __init__(self, image_dir: str = ""):
        """
        Args:
            image_dir: 이미지 저장 디렉토리 경로. 비어있으면 로컬 저장 안 함.
        """
        self.image_dir = image_dir
        if image_dir:
            Path(image_dir).mkdir(parents=True, exist_ok=True)

        # 스타일 정의
        self.header_font = Font(name="맑은 고딕", size=11, bold=True)
        self.header_fill = PatternFill(
            start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.data_font = Font(name="맑은 고딕", size=11)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

        # 테두리 스타일
        self.thin_side = Side(style="thin", color="000000")
        self.medium_side = Side(style="medium", color="000000")

    def create_workbook(self, itinerary_days: List[ItineraryDay]) -> openpyxl.Workbook:
        """
        일정 데이터로부터 엑셀 워크북 생성

        Args:
            itinerary_days: 일자별 일정 리스트

        Returns:
            openpyxl.Workbook: 생성된 워크북 객체
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "여행일정"

        # 컬럼 너비 설정
        self._set_column_widths(ws)

        # 헤더 작성
        self._write_header(ws)

        # 이미지 추가 정보 저장
        images_to_add = []

        # 데이터 작성
        current_row = 2
        for day in itinerary_days:
            start_row = current_row
            current_row = self._write_day(ws, day, start_row, images_to_add)
            # 일자 컬럼 병합
            if current_row > start_row:
                ws.merge_cells(f"A{start_row}:A{current_row - 1}")

        # 이미지 추가 (URL에서 다운로드)
        for img_info in images_to_add:
            try:
                self._add_image_from_url(ws, img_info["row"], img_info["col"], img_info["url"], img_info.get("label", ""))
            except Exception:
                # 이미지 다운로드 실패 시 URL 텍스트로 대체
                ws.cell(row=img_info["row"], column=img_info["col"]).value = img_info["url"]

        return wb

    def _add_image_from_url(self, ws, row: int, col: int, url: str, label: str = ""):
        """URL에서 이미지를 다운로드하여 셀에 추가 + 로컬 저장"""
        if not url:
            return

        try:
            # 이미지 다운로드
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as response:
                raw_bytes = response.read()

            # 로컬 저장
            if self.image_dir:
                self._save_image_locally(raw_bytes, url, label)

            img_data = io.BytesIO(raw_bytes)

            # openpyxl 이미지 객체 생성
            img = XLImage(img_data)

            # 이미지 크기 조정 (셀에 맞게)
            img.width = 100
            img.height = 75

            # 셀 위치 계산
            cell_ref = f"{get_column_letter(col)}{row}"
            ws.add_image(img, cell_ref)

        except Exception:
            # 실패 시 조용히 무시 (URL 텍스트로 대체됨)
            raise

    def _save_image_locally(self, raw_bytes: bytes, url: str, label: str = ""):
        """다운로드한 이미지를 로컬 디렉토리에 저장"""
        # URL에서 파일명 추출
        filename = url.rstrip("/").split("/")[-1]
        # label이 있으면 파일명 앞에 붙여서 식별 용이하게
        if label:
            safe_label = re.sub(r'[\\/*?:"<>|]', '_', label)[:40].strip('_. ')
            ext = Path(filename).suffix or ".jpg"
            filename = f"{safe_label}{ext}"

        filepath = Path(self.image_dir) / filename

        # 동일 파일명 충돌 방지
        if filepath.exists():
            stem = filepath.stem
            ext = filepath.suffix
            i = 1
            while filepath.exists():
                filepath = Path(self.image_dir) / f"{stem}_{i}{ext}"
                i += 1

        filepath.write_bytes(raw_bytes)

    def _set_column_widths(self, ws):
        """컬럼 너비 설정"""
        ws.column_dimensions["A"].width = 10  # 일자
        ws.column_dimensions["B"].width = 15  # 지역
        ws.column_dimensions["C"].width = 20  # 관광지
        ws.column_dimensions["D"].width = 15  # 이미지
        ws.column_dimensions["E"].width = 60  # 관광지 설명

    def _write_header(self, ws):
        """헤더 행 작성"""
        headers = ["일자", "지역", "관광지", "이미지", "관광지 설명"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = Border(
                left=self.medium_side if col_idx == 1 else self.thin_side,
                right=self.medium_side if col_idx == 5 else self.thin_side,
                top=self.medium_side,
                bottom=self.thin_side,
            )

        ws.row_dimensions[1].height = 30

    def _write_day(
        self, ws, day: ItineraryDay, start_row: int, images_to_add: list
    ) -> int:
        """
        일자별 일정 작성

        Args:
            ws: 워크시트
            day: 일자 데이터
            start_row: 시작 행 번호
            images_to_add: 이미지 추가 정보 리스트 (나중에 일괄 추가)

        Returns:
            int: 다음 행 번호
        """
        current_row = start_row

        for idx, item in enumerate(day.items):
            # 첫 번째 항목만 일자 표시
            if idx == 0:
                ws.cell(row=current_row, column=1).value = f"{day.day_number}일차"

            # 지역
            region = item.region_from if item.region_from else item.region_to
            ws.cell(row=current_row, column=2).value = region

            # 관광지 이름
            ws.cell(row=current_row, column=3).value = item.schedule_title

            # 이미지 URL 저장 (나중에 일괄 추가)
            if item.image_url:
                images_to_add.append({
                    "row": current_row,
                    "col": 4,
                    "url": item.image_url,
                    "label": item.schedule_title,
                })

            # 관광지 설명
            ws.cell(row=current_row, column=5).value = item.description

            # 스타일 적용
            self._apply_cell_styles(ws, current_row)

            # 행 높이 설정 (이미지가 있으면 더 크게)
            if item.image_url:
                ws.row_dimensions[current_row].height = 80
            else:
                ws.row_dimensions[current_row].height = 60

            current_row += 1

        return current_row

    def _apply_cell_styles(self, ws, row: int):
        """셀 스타일 적용"""
        for col_idx in range(1, 6):  # A~E (5개 컬럼)
            cell = ws.cell(row=row, column=col_idx)
            cell.font = self.data_font

            # 정렬
            if col_idx in [1, 2, 3, 4]:  # 일자, 지역, 관광지, 이미지
                cell.alignment = self.center_alignment
            elif col_idx == 5:  # 관광지 설명
                cell.alignment = self.left_alignment

            # 테두리
            left_border = self.medium_side if col_idx == 1 else self.thin_side
            right_border = self.medium_side if col_idx == 5 else self.thin_side

            cell.border = Border(
                left=left_border,
                right=right_border,
                top=self.thin_side,
                bottom=self.thin_side,
            )

    def generate_excel_bytes(self, itinerary_days: List[ItineraryDay]) -> bytes:
        """
        엑셀 파일을 바이트로 생성 (API 응답용)

        Args:
            itinerary_days: 일자별 일정 리스트

        Returns:
            bytes: 엑셀 파일 바이트 데이터
        """
        wb = self.create_workbook(itinerary_days)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def save_excel(self, itinerary_days: List[ItineraryDay], filepath: str):
        """
        엑셀 파일을 디스크에 저장

        Args:
            itinerary_days: 일자별 일정 리스트
            filepath: 저장 경로
        """
        wb = self.create_workbook(itinerary_days)
        wb.save(filepath)
