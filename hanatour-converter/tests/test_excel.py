"""
Excel Generator Test

엑셀 생성 서비스 테스트
"""

import os
from datetime import time
from pathlib import Path

import openpyxl
import pytest

from app.services.excel import ExcelGenerator, ItineraryDay, ItineraryItem


@pytest.fixture
def excel_generator():
    """ExcelGenerator 인스턴스 픽스처"""
    return ExcelGenerator()


@pytest.fixture
def sample_itinerary():
    """샘플 일정 데이터"""
    day1_items = [
        ItineraryItem(
            region_from="인천",
            transport="VN431",
            time=time(10, 10, 0),
            description="인천 국제공항 출발",
            meal="조:기내식",
        ),
        ItineraryItem(
            region_to="다낭",
            time=time(13, 10, 0),
            description="다낭 국제공항 도착 및 입국 수속",
            meal="석:현지정식",
        ),
        ItineraryItem(
            description=" - 공항에서 가이드 미팅 및 한국어 가이드 합류",
        ),
        ItineraryItem(
            transport="전용차량",
            description="전용차량 탑승 및 시내로 이동 후 중식",
        ),
        ItineraryItem(
            description="▶ 다낭 현지의 분위기를 느낄 수 있는 시장 방문",
        ),
        ItineraryItem(
            description="▶ 다낭 핫플레이스 카페거리 방문 및 커피 타임",
        ),
        ItineraryItem(
            description="▶ 베트남 전통 발 마사지로 여행 피로를 풀며 90분 체험",
        ),
        ItineraryItem(
            description="저녁 식사 후 호텔 이동 및 휴식",
        ),
        ItineraryItem(
            description="HOTEL : 빈펄 리조트 다낭 리버프론트  [5성급]",
        ),
    ]

    day2_items = [
        ItineraryItem(
            region_from="다낭",
            transport="전용차량",
            time=time(8, 0, 0),
            description="호텔 조식 후 가이드 미팅",
            meal="조:호텔식",
        ),
        ItineraryItem(
            description="▶ 다낭 대표 관광지 방문",
            meal="중:현지정식",
        ),
        ItineraryItem(
            description=" - 가이드와 함께 유명 사원, 시장 투어",
            meal="석:한식 불고기",
        ),
        ItineraryItem(
            description=" - 중식은 유명한 해산물 레스토랑",
        ),
        ItineraryItem(
            description="오후 다낭 시내로 이동",
        ),
        ItineraryItem(
            description="▶ 포토제닉한 장소에서 인생샷 촬영 90분 체험",
        ),
        ItineraryItem(
            description="저녁 식사 후 호텔 이동 및 휴식",
        ),
        ItineraryItem(
            description="HOTEL : 빈펄 리조트 다낭 리버프론트  [5성급]",
        ),
    ]

    day3_items = [
        ItineraryItem(
            region_from="다낭",
            transport="전용차량",
            time=time(9, 0, 0),
            description="호텔 조식 후 가이드 미팅",
            meal="조:호텔식",
        ),
        ItineraryItem(
            description="▶ 전 세계가 주목하는 테마파크 선셋쇼 관람",
            meal="중:뷔페식",
        ),
        ItineraryItem(
            description=" - 다양한 어트랙션 및 공연 관람",
        ),
        ItineraryItem(
            description="다낭 공항으로 이동",
            meal="석:기내식",
        ),
        ItineraryItem(
            region_to="인천",
            transport="VN432",
            time=time(22, 30, 0),
            description="다낭 국제공항 출발",
        ),
    ]

    return [
        ItineraryDay(day_number=1, items=day1_items),
        ItineraryDay(day_number=2, items=day2_items),
        ItineraryDay(day_number=3, items=day3_items),
    ]


def test_excel_generator_initialization(excel_generator):
    """ExcelGenerator 초기화 테스트"""
    assert excel_generator is not None
    assert excel_generator.header_font is not None
    assert excel_generator.header_fill is not None
    assert excel_generator.data_font is not None


def test_create_workbook(excel_generator, sample_itinerary):
    """워크북 생성 테스트"""
    wb = excel_generator.create_workbook(sample_itinerary)

    assert wb is not None
    assert wb.active.title == "여행일정"

    ws = wb.active

    # 헤더 확인
    assert ws.cell(1, 1).value == "일자"
    assert ws.cell(1, 2).value == "지역"
    assert ws.cell(1, 3).value == "교통편"
    assert ws.cell(1, 4).value == "시간"
    assert ws.cell(1, 5).value == "일정"
    assert ws.cell(1, 8).value == "식사"

    # 데이터 확인 (1일차 첫 행)
    assert ws.cell(2, 1).value == "1일차"
    assert ws.cell(2, 2).value == "인천"
    assert ws.cell(2, 3).value == "VN431"
    assert ws.cell(2, 5).value == "인천 국제공항 출발"
    assert ws.cell(2, 8).value == "조:기내식"

    # 컬럼 너비 확인
    assert ws.column_dimensions["A"].width == 13.75
    assert ws.column_dimensions["B"].width == 11.75
    assert ws.column_dimensions["H"].width == 19.25


def test_merged_cells(excel_generator, sample_itinerary):
    """병합 셀 테스트"""
    wb = excel_generator.create_workbook(sample_itinerary)
    ws = wb.active

    # 헤더 일정 컬럼 병합 확인
    assert "E1:G1" in [str(merged) for merged in ws.merged_cells.ranges]

    # 데이터 행 일정 컬럼 병합 확인
    assert "E2:G2" in [str(merged) for merged in ws.merged_cells.ranges]
    assert "E3:G3" in [str(merged) for merged in ws.merged_cells.ranges]


def test_cell_styles(excel_generator, sample_itinerary):
    """셀 스타일 테스트"""
    wb = excel_generator.create_workbook(sample_itinerary)
    ws = wb.active

    # 헤더 스타일 확인
    header_cell = ws.cell(1, 1)
    assert header_cell.font.bold is True
    assert header_cell.font.name == "맑은 고딕"
    assert header_cell.alignment.horizontal == "center"

    # 데이터 셀 스타일 확인
    data_cell = ws.cell(2, 1)
    assert data_cell.font.name == "맑은 고딕"
    assert data_cell.alignment.horizontal == "center"

    # 일정 셀 정렬 확인
    itinerary_cell = ws.cell(2, 5)
    assert itinerary_cell.alignment.horizontal == "left"
    assert itinerary_cell.alignment.wrap_text is True


def test_generate_excel_bytes(excel_generator, sample_itinerary):
    """바이트 생성 테스트"""
    excel_bytes = excel_generator.generate_excel_bytes(sample_itinerary)

    assert excel_bytes is not None
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0

    # 바이트가 유효한 엑셀 파일인지 확인
    import io

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert wb.active.title == "여행일정"


def test_save_excel(excel_generator, sample_itinerary, tmp_path):
    """파일 저장 테스트"""
    output_path = tmp_path / "test_output.xlsx"
    excel_generator.save_excel(sample_itinerary, str(output_path))

    assert output_path.exists()

    # 저장된 파일 읽기
    wb = openpyxl.load_workbook(output_path)
    assert wb.active.title == "여행일정"
    assert wb.active.cell(1, 1).value == "일자"


def test_save_excel_to_downloads(excel_generator, sample_itinerary):
    """downloads 폴더에 샘플 엑셀 생성 테스트"""
    downloads_dir = Path(__file__).parent.parent / "downloads"
    downloads_dir.mkdir(exist_ok=True)

    output_path = downloads_dir / "sample_itinerary.xlsx"
    excel_generator.save_excel(sample_itinerary, str(output_path))

    assert output_path.exists()
    assert output_path.stat().st_size > 0

    print(f"\n✓ 샘플 엑셀 파일 생성 완료: {output_path}")


def test_time_format(excel_generator):
    """시간 포맷 테스트"""
    items = [
        ItineraryItem(
            region_from="서울",
            time=time(14, 30, 0),
            description="서울 출발",
        )
    ]
    day = ItineraryDay(day_number=1, items=items)

    wb = excel_generator.create_workbook([day])
    ws = wb.active

    time_cell = ws.cell(2, 4)
    assert time_cell.value is not None
    assert time_cell.number_format == "hh:mm:ss"


def test_empty_itinerary(excel_generator):
    """빈 일정 테스트"""
    wb = excel_generator.create_workbook([])
    ws = wb.active

    # 헤더만 존재해야 함
    assert ws.cell(1, 1).value == "일자"
    assert ws.cell(2, 1).value is None


def test_single_day_multiple_items(excel_generator):
    """하루에 여러 항목 테스트"""
    items = [
        ItineraryItem(description=f"일정 항목 {i+1}") for i in range(5)
    ]
    day = ItineraryDay(day_number=1, items=items)

    wb = excel_generator.create_workbook([day])
    ws = wb.active

    # 일자 컬럼 병합 확인
    assert "A2:A6" in [str(merged) for merged in ws.merged_cells.ranges]

    # 각 항목 확인
    for i in range(5):
        assert ws.cell(2 + i, 5).value == f"일정 항목 {i+1}"
